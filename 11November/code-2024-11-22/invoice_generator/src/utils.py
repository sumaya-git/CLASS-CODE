"""Contains helper functions"""

import calendar
from datetime import date

from openpyxl import load_workbook


def get_company_info() -> dict[str, str | list[str]]:
    # TODO: Extract these info from either a csv or json file
    return {
        "company_name": "DCI-Digital Career Institute GmbH",
        "company_address": ["Vulkanstr. 1", "10367", "Berlin"],
    }


def get_user_info():
    return {
        "user_name": "Cassidy Potter",
        "user_profession": "Backend Developer",
        "user_email": "cassidy@example.com",
        "po_number": "PO-598",
        "rate_per_hour": 20,
        "default_class": "P23EO4",
    }


def period_of_service():
    service_year = date.today().year
    service_month = date.today().month
    _, days_count = calendar.monthrange(service_year, service_month)
    service_period_start = date(service_year, service_month, 1)
    service_period_end = date(service_year, service_month, days_count)

    formated_period_start = "/".join(str(service_period_start).split("-")[::-1])
    formated_period_end = "/".join(str(service_period_end).split("-")[::-1])

    # formatted_start = service_period_start.strftime("%d/%m/%Y")
    # formatted_end = service_period_end.strftime("%d/%m/%Y")

    return formated_period_start, formated_period_end


def get_invoice_header_info() -> dict[str, str | tuple[str, str]]:
    return {
        "invoice_date": date.today().strftime("%d %b %Y"),
        "period_of_service": period_of_service(),
        "invoice_number": f"010 {date.today().year}",
    }


def extract_excel_file(file_path: str):
    # FIXME: Improve on this function

    wb = load_workbook(filename=file_path)

    sheet = wb.active

    all_rows = []

    for row in range(1, sheet.max_row + 1):  # loop row
        sheet_row = []
        for col in range(1, sheet.max_column + 1):
            sheet_row.append(sheet.cell(column=col, row=row).value)
        all_rows.append(sheet_row)

    return all_rows